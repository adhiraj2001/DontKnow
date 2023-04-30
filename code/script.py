import openpyxl
from cognitive_attributes import attributes as cognitive_attributes
from subject_attributes import attributes as subject_attributes
from blood_test_attributes import attributes as blood_test_attributes

#print(cognitive_attributes)
workbook = openpyxl.load_workbook("./DFS_course_Sample_data.xlsx")

worksheet = workbook.active

num_rows = worksheet.max_row
num_cols = worksheet.max_column
key_index_map = {}
index_key_map = {}
row_count = 0

T_Subject = []
T_Cognitive_Test = []
T_Blood_Test = []
T_CamCan = []
# Or iterate over rows and columns to retrieve data from the worksheet
for row in worksheet.iter_rows(min_row=1, max_row=num_rows, min_col=1, max_col=num_cols):
    curr_col = 0
    
    subject_row = {}
    cognitive_row = {}
    blood_row = {}


    for cell in row:
        if row_count == 0:
            key_index_map[cell.value] = curr_col
            index_key_map[curr_col] = cell.value
        else:
            curr_col_name = index_key_map[curr_col]
            if curr_col_name in subject_attributes:
                subject_row[curr_col_name] = cell.value

            if curr_col_name in cognitive_attributes:
                cognitive_row[curr_col_name] = cell.value

            if curr_col_name in blood_test_attributes:
                blood_row[curr_col_name] = cell.value
        curr_col += 1
    if row_count > 0:
        T_Cognitive_Test.append(cognitive_row)
        T_Blood_Test.append(blood_row)
        T_Subject.append(subject_row)

    row_count += 1





